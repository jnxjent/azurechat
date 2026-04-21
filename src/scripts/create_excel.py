"""
create_excel.py  –  openpyxl ベースの Excel 新規作成スクリプト
CLI: python create_excel.py --output <path> --plan <json_path>
stdout: JSON {"sheets": N, "totalRows": N}

plan JSON schema:
{
  "sheets": [
    {
      "name": "Sheet1",
      "rows": [
        ["", "野元", "山本", "鈴木"],
        ["点数", 100, 60, 40]
      ],
      "headerRowIndex": 0,   // この行をヘッダーとして太字＋背景色にする（省略可）
      "autoWidth": true       // 列幅を自動調整する（省略時 true）
    }
  ]
}
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF4472C4")  # Excel blue
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
ALT_FILL = PatternFill(fill_type="solid", fgColor="FFDCE6F1")  # light blue alt row


def cast_value(raw: Any) -> Any:
    """文字列として渡された数値を int / float に変換する。"""
    if isinstance(raw, (int, float)):
        return raw
    if isinstance(raw, str):
        s = raw.strip()
        if s == "":
            return s
        try:
            return int(s)
        except ValueError:
            pass
        try:
            return float(s)
        except ValueError:
            pass
    return raw


def auto_fit_columns(ws) -> None:
    """各列の最大文字数を基に列幅を設定する。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                # CJK文字は2幅として計算
                cjk_extra = sum(
                    1
                    for ch in str(cell.value or "")
                    if "\u3000" <= ch <= "\u9fff" or "\u3040" <= ch <= "\u30ff" or "\uff00" <= ch <= "\uffef"
                )
                cell_len += cjk_extra
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)


def build_sheet(wb: openpyxl.Workbook, sheet_def: dict) -> int:
    """シート定義からワークシートを構築。書き込んだ行数を返す。"""
    name = str(sheet_def.get("name") or "Sheet1").strip()[:31]  # Excel max 31 chars
    rows = sheet_def.get("rows") or []
    header_row_index = sheet_def.get("headerRowIndex")
    do_auto_width = sheet_def.get("autoWidth", True)

    ws = wb.create_sheet(title=name)
    written = 0

    for ri, row_data in enumerate(rows):
        for ci, cell_val in enumerate(row_data):
            cell = ws.cell(row=ri + 1, column=ci + 1, value=cast_value(cell_val))

            is_header = (header_row_index is not None and ri == header_row_index)

            if is_header:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            else:
                cell.border = THIN_BORDER
                # 偶数行に薄い背景
                if header_row_index is not None and ri % 2 == 0:
                    cell.fill = ALT_FILL

        written += 1

    if do_auto_width:
        auto_fit_columns(ws)

    return written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--plan", required=True)
    args = parser.parse_args()

    output_path = Path(args.output)
    plan: dict[str, Any] = json.loads(Path(args.plan).read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトの空シートを削除

    total_sheets = 0
    total_rows = 0

    for sheet_def in plan.get("sheets") or []:
        n = build_sheet(wb, sheet_def)
        total_rows += n
        total_sheets += 1

    if total_sheets == 0:
        # sheets が空でも最低1シートは出力
        ws = wb.create_sheet(title="Sheet1")
        total_sheets = 1

    wb.save(str(output_path))
    print(json.dumps({"sheets": total_sheets, "totalRows": total_rows}))


if __name__ == "__main__":
    main()
