"""
edit_excel.py  –  openpyxl ベースの Excel 編集スクリプト
CLI: python edit_excel.py --input <path> --output <path> --plan <json_path>
stdout: JSON {"changedSheets": N, "totalSheets": N}
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple


def _replace_stdev_with_sqrt_sumproduct(formula: str) -> str:
    """STDEVP(range) / STDEV.P(range) を全Excelバージョン互換の
    SQRT(SUMPRODUCT((range-AVERAGE(range))^2)/COUNT(range)) に置換する。
    range 部分は括弧の深さを追って正確に取り出す。"""
    pattern = re.compile(r"\bSTDEV(?:\.P|P)\s*\(", re.IGNORECASE)
    result = []
    i = 0
    while i < len(formula):
        m = pattern.search(formula, i)
        if not m:
            result.append(formula[i:])
            break
        result.append(formula[i:m.start()])
        depth = 1
        j = m.end()
        while j < len(formula) and depth > 0:
            if formula[j] == "(":
                depth += 1
            elif formula[j] == ")":
                depth -= 1
            j += 1
        rng = formula[m.end():j - 1]
        avg = f"AVERAGE({rng})"
        replacement = f"SQRT(SUMPRODUCT(({rng}-{avg})^2)/COUNT({rng}))"
        result.append(replacement)
        i = j
    return "".join(result)


def sanitize_formula(value: Any) -> Any:
    """LLM が生成した Excel 式を正規化する。
    - 全角・半角 @ 演算子の除去（例: @STDEV.P → STDEV.P）
    - STDEV.P / STDEVP → SQRT(SUMPRODUCT(...)) 変換（全Excelバージョン互換）
    """
    if not isinstance(value, str) or not value.startswith("="):
        return value
    # 全角 @ を半角に統一してから除去
    normalized = value.replace("＠", "@")
    normalized = re.sub(r"@(?=[A-Za-z_])", "", normalized)
    # STDEV.P / STDEVP を SQRT(SUMPRODUCT) に変換（互換性確保）
    normalized = _replace_stdev_with_sqrt_sumproduct(normalized)
    if normalized != value:
        print(
            f"[edit_excel] sanitize_formula: {value!r} -> {normalized!r}",
            file=sys.stderr,
        )
    return normalized


def parse_hex_color(value: str | None) -> str | None:
    """6桁 RRGGBB → openpyxl 用 ARGB (FF + RRGGBB)。無効な値は None を返す。"""
    if not value:
        return None
    normalized = str(value).replace("#", "").strip().upper()
    if len(normalized) == 6 and all(ch in "0123456789ABCDEF" for ch in normalized):
        return "FF" + normalized
    return None


def apply_sheet_edits(ws, edits: dict) -> int:
    """setCells / replaceText をシートに適用。変更したセル数を返す。"""
    changed = 0

    for cell_edit in edits.get("setCells") or []:
        address = str(cell_edit.get("address") or "").strip()
        value = cell_edit.get("value")
        if address and value is not None:
            ws[address] = sanitize_formula(value)
            changed += 1

    replacements = edits.get("replaceText") or []
    if replacements:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                original = cell.value
                new_val = original
                for rep in replacements:
                    find = rep.get("find", "")
                    replace = rep.get("replace", "")
                    if find and find in new_val:
                        new_val = new_val.replace(find, replace)
                if new_val != original:
                    cell.value = new_val
                    changed += 1

    return changed


def apply_format_edits(ws, fmt: dict) -> int:
    """1つの formatEdit エントリを適用。変更したセル数を返す。"""
    range_str = str(fmt.get("range") or "").strip()
    if not range_str:
        return 0

    bold = fmt.get("bold")
    font_color = parse_hex_color(fmt.get("fontColor"))
    fill_color = parse_hex_color(fmt.get("fillColor"))

    if bold is None and font_color is None and fill_color is None:
        return 0

    changed = 0
    try:
        cell_range = ws[range_str]
        # 単一セルの場合は list にする
        if not isinstance(cell_range, (list, tuple)):
            cell_range = [[cell_range]]
        for row in cell_range:
            row_cells = row if isinstance(row, (list, tuple)) else [row]
            for cell in row_cells:
                if bold is not None or font_color is not None:
                    f = cell.font if cell.font else Font()
                    kwargs: dict[str, Any] = {}
                    if bold is not None:
                        kwargs["bold"] = bold
                    if font_color is not None:
                        kwargs["color"] = font_color
                    cell.font = f.copy(**kwargs)
                if fill_color is not None:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                changed += 1
    except Exception as e:
        print(f"[edit_excel] format range '{range_str}' failed: {e}", file=sys.stderr)

    return changed


def load_workbook_any(input_path: Path) -> openpyxl.Workbook:
    """xlsx / xlsm は openpyxl で直接ロード。xls は xlrd 経由で変換。"""
    suffix = input_path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        keep_vba = suffix == ".xlsm"
        return openpyxl.load_workbook(str(input_path), keep_vba=keep_vba)

    if suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError:
            raise RuntimeError(
                "xlrd が未インストールのため .xls ファイルを読み込めません。"
                "startup.sh で pip install xlrd を実行してください。"
            )
        book = xlrd.open_workbook(str(input_path))
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除
        for sheet_idx in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_idx)
            ws = wb.create_sheet(title=sheet.name)
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    ctype = sheet.cell_type(row_idx, col_idx)
                    if ctype == 0:  # empty
                        continue
                    ws.cell(
                        row=row_idx + 1,
                        column=col_idx + 1,
                        value=sheet.cell_value(row_idx, col_idx),
                    )
        return wb

    raise ValueError(f"Unsupported file format: {suffix}")


def resolve_sheet(wb: openpyxl.Workbook, sheet_name: str):
    """シート名で検索。完全一致 → 前方一致の順で返す。見つからなければ None。"""
    if not sheet_name or sheet_name == "*":
        return None  # 全シート対象を呼び元で処理
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    match = next(
        (s for s in wb.sheetnames if s.lower().startswith(sheet_name.lower())), None
    )
    return wb[match] if match else None


def resolve_column_index(ws, col_spec: str) -> int | None:
    """列指定を列番号（1始まり）に解決。
    - "A"〜"XFD" 形式はそのまま変換
    - ヘッダー名の場合は1行目を走査して列番号を返す
    """
    if not col_spec:
        return None
    col_spec = col_spec.strip()
    # アルファベットのみなら列記号として解釈
    if col_spec.isalpha():
        try:
            return column_index_from_string(col_spec)
        except Exception:
            pass
    # ヘッダー名として1行目を検索
    for cell in ws[1]:
        if str(cell.value or "").strip() == col_spec:
            return cell.column
    print(f"[edit_excel] resolve_column_index: '{col_spec}' not found in row 1", file=sys.stderr)
    return None


def apply_border_edits(ws, border_edit: dict) -> int:
    """指定範囲に罫線を適用する。
    edges: "all" | "outer" | "inner" | "top" | "bottom" | "left" | "right"
    style: "thin" | "medium" | "thick" | "hair" | "dashed" (デフォルト "thin")
    """
    range_str = str(border_edit.get("range") or "").strip()
    if not range_str:
        return 0
    style = str(border_edit.get("style") or "thin")
    edges = str(border_edit.get("edges") or "all").lower()
    side = Side(style=style)
    no_side = Side(style=None)

    try:
        cell_range = ws[range_str]
        if not isinstance(cell_range, (list, tuple)):
            cell_range = [[cell_range]]
        # 範囲の行数・列数を把握
        rows = cell_range
        max_r = len(rows) - 1
        changed = 0
        for ri, row in enumerate(rows):
            row_cells = row if isinstance(row, (list, tuple)) else [row]
            max_c = len(row_cells) - 1
            for ci, cell in enumerate(row_cells):
                b = cell.border.copy() if cell.border else Border()
                top = b.top
                bottom = b.bottom
                left = b.left
                right = b.right

                is_top_edge = ri == 0
                is_bottom_edge = ri == max_r
                is_left_edge = ci == 0
                is_right_edge = ci == max_c

                if edges == "all":
                    top = bottom = left = right = side
                elif edges == "outer":
                    if is_top_edge:    top = side
                    if is_bottom_edge: bottom = side
                    if is_left_edge:   left = side
                    if is_right_edge:  right = side
                elif edges == "inner":
                    if not is_top_edge:    top = side
                    if not is_bottom_edge: bottom = side
                    if not is_left_edge:   left = side
                    if not is_right_edge:  right = side
                elif edges == "top":    top = side
                elif edges == "bottom": bottom = side
                elif edges == "left":   left = side
                elif edges == "right":  right = side

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)
                changed += 1
        return changed
    except Exception as e:
        print(f"[edit_excel] border range '{range_str}' failed: {e}", file=sys.stderr)
        return 0


def apply_copy_row_color(ws, target_col_spec: str, reference_col_spec: str, start_row: int = 2) -> int:
    """reference 列の各行の背景色を target 列の同じ行にコピーする。
    col_spec は列記号（"A"）またはヘッダー名（"対応"）のどちらでもよい。
    変更した行数を返す。
    """
    tgt_idx = resolve_column_index(ws, target_col_spec)
    ref_idx = resolve_column_index(ws, reference_col_spec)

    if tgt_idx is None or ref_idx is None:
        print(
            f"[edit_excel] copy_row_color: could not resolve columns "
            f"target='{target_col_spec}' ref='{reference_col_spec}'",
            file=sys.stderr,
        )
        return 0

    changed = 0
    max_row = ws.max_row
    for row_num in range(start_row, max_row + 1):
        ref_cell = ws.cell(row=row_num, column=ref_idx)
        tgt_cell = ws.cell(row=row_num, column=tgt_idx)

        ref_fill = ref_cell.fill
        if ref_fill and ref_fill.fill_type not in (None, "none", ""):
            # PatternFill をコピー（fgColor / bgColor ともにそのまま渡す）
            tgt_cell.fill = PatternFill(
                fill_type=ref_fill.fill_type,
                fgColor=ref_fill.fgColor.rgb if ref_fill.fgColor else "00000000",
                bgColor=ref_fill.bgColor.rgb if ref_fill.bgColor else "00000000",
            )
            changed += 1

    return changed


def apply_chart_edits(ws, chart_edit: dict, output_path: "Path" = None) -> bool:
    """matplotlib でグラフ画像を生成し openpyxl Image としてシートに挿入する。"""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.font_manager as fm
        from openpyxl.drawing.image import Image as XLImage
    except ImportError as e:
        print(f"[edit_excel] chart: matplotlib/pillow not available ({e}). Run: pip install matplotlib pillow", file=sys.stderr)
        return False

    chart_type = str(chart_edit.get("chartType") or "line").lower()
    title = str(chart_edit.get("title") or "")
    x_col_spec = str(chart_edit.get("xColumn") or "")
    raw_y = chart_edit.get("yColumns") or []
    y_col_specs: list[str] = [raw_y] if isinstance(raw_y, str) else list(raw_y)
    x_label = str(chart_edit.get("xLabel") or "")
    y_label = str(chart_edit.get("yLabel") or "")
    insert_cell = str(chart_edit.get("insertCell") or "E2")

    def chart_float(name: str) -> float | None:
        raw = chart_edit.get(name)
        if raw in (None, ""):
            return None
        try:
            return float(str(raw).replace(",", ""))
        except (TypeError, ValueError):
            print(f"[edit_excel] chart: invalid {name}={raw!r}; ignoring", file=sys.stderr)
            return None

    _JP_COLOR_MAP: dict[str, str] = {
        "赤": "red", "青": "blue", "緑": "green", "黄": "yellow",
        "橙": "orange", "オレンジ": "orange", "紫": "purple",
        "黒": "black", "白": "white", "灰": "gray", "グレー": "gray",
        "ピンク": "pink", "水色": "cyan", "シアン": "cyan",
        "茶": "brown", "茶色": "brown",
    }

    def _resolve_color(c: str) -> str | None:
        import matplotlib.colors as _mcolors
        c = str(c).strip()
        # 日本語色名マップ
        if c in _JP_COLOR_MAP:
            c = _JP_COLOR_MAP[c]
        if _mcolors.is_color_like(c):
            return c
        print(f"[edit_excel] chart: invalid color '{c}' — ignored", file=sys.stderr)
        return None

    raw_colors = chart_edit.get("seriesColors") or []
    raw_list: list[str] = [raw_colors] if isinstance(raw_colors, str) else [str(c) for c in raw_colors if c]
    series_colors: list[str | None] = [_resolve_color(c) for c in raw_list]

    x_idx = resolve_column_index(ws, x_col_spec) if x_col_spec else None
    y_divisor = chart_float("yDivisor") or 1.0
    if y_divisor == 0:
        y_divisor = 1.0

    y_min = chart_float("yMin")
    y_max = chart_float("yMax")
    y_tick_step = chart_float("yTickStep")
    y_tick_format = str(chart_edit.get("yTickFormat") or "auto").lower()

    y_indices: list[int] = []
    y_headers: list[str] = []
    for y_spec in y_col_specs:
        idx = resolve_column_index(ws, y_spec)
        if idx is not None:
            y_indices.append(idx)
            y_headers.append(str(ws.cell(row=1, column=idx).value or y_spec))

    if not y_indices:
        print(f"[edit_excel] chart: no valid yColumns found", file=sys.stderr)
        return False

    # yDivisor != 1 なのに yLabel に単位がなければ自動補完（y_headers確定後に実行）
    _unit_map = {1000.0: "千円", 10000.0: "万円", 1000000.0: "百万円", 100000000.0: "億円"}
    _auto_unit = _unit_map.get(y_divisor)
    if _auto_unit:
        _base_label = y_label if y_label else (y_headers[0] if y_headers else "値")
        if f"（{_auto_unit}）" not in _base_label and f"({_auto_unit})" not in _base_label:
            y_label = f"{_base_label}（{_auto_unit}）"
            print(f"[edit_excel] chart: auto-set yLabel: '{y_label}'", file=sys.stderr)

    # データ読み取り（2行目以降）Y値がある行だけ採用、X空なら連番ラベル
    x_labels: list[str] = []
    y_data_sets: list[list[float]] = [[] for _ in y_indices]
    for row_num in range(2, ws.max_row + 1):
        row_y: list[float] = []
        has_y = False
        for y_idx in y_indices:
            raw = ws.cell(row=row_num, column=y_idx).value
            try:
                val = float(raw) if raw not in (None, "") else None
            except (TypeError, ValueError):
                val = None  # 文字列（ヘッダー等）はスキップ
            if val is not None:
                has_y = True
                row_y.append(val / y_divisor)
            else:
                row_y.append(0.0)
        if not has_y:
            continue
        x_raw = ws.cell(row=row_num, column=x_idx).value if x_idx is not None else None
        x_labels.append(str(x_raw).strip() if x_raw not in (None, "") else str(len(x_labels) + 1))
        for i, v in enumerate(row_y):
            y_data_sets[i].append(v)

    if not x_labels:
        print(f"[edit_excel] chart: no data rows found (sheet has {ws.max_row} rows, yColumns={y_col_specs})", file=sys.stderr)
        return False

    print(f"[edit_excel] chart: {len(x_labels)} rows, x={x_labels[:5]}, y[0]={y_data_sets[0][:5]}", file=sys.stderr)

    # 日本語フォント設定
    # 1. リポジトリ同梱フォント (public/fonts/NotoSansJP-Regular.ttf) を優先ロード
    _script_dir = Path(__file__).resolve().parent
    _bundled_font = None
    for _candidate in [
        _script_dir / ".." / ".." / "public" / "fonts" / "NotoSansJP-Regular.ttf",
        _script_dir / ".." / "public" / "fonts" / "NotoSansJP-Regular.ttf",
        Path("/home/site/wwwroot/public/fonts/NotoSansJP-Regular.ttf"),
    ]:
        _resolved = _candidate.resolve()
        if _resolved.exists():
            fm.fontManager.addfont(str(_resolved))
            _bundled_font = fm.FontProperties(fname=str(_resolved)).get_name()
            print(f"[edit_excel] chart: loaded bundled font '{_bundled_font}': {_resolved}", file=sys.stderr)
            break

    # 2. システムフォントフォールバック
    if _bundled_font is None:
        _jp_candidates = [
            "IPAexGothic", "IPAPGothic", "Noto Sans CJK JP", "Noto Sans JP",
            "MS Gothic", "Yu Gothic", "Hiragino Sans", "TakaoPGothic",
        ]
        _available = {f.name for f in fm.fontManager.ttflist}
        _bundled_font = next((f for f in _jp_candidates if f in _available), None)

    # font.family="sans-serif" + font.sans-serif リスト先頭に追加（最も確実な指定方法）
    if _bundled_font:
        plt.rcParams["font.family"] = "sans-serif"
        plt.rcParams["font.sans-serif"] = [_bundled_font] + list(plt.rcParams.get("font.sans-serif", []))
    else:
        plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["axes.unicode_minus"] = False

    fig, ax = plt.subplots(figsize=(8, 5))
    x_pos = list(range(len(x_labels)))

    if chart_type == "bar":
        n = max(len(y_indices), 1)
        width = 0.7 / n
        for i, (y_vals, label) in enumerate(zip(y_data_sets, y_headers)):
            offsets = [p + (i - (n - 1) / 2) * width for p in x_pos]
            color = series_colors[i] if i < len(series_colors) else None
            kwargs: dict = {"width": width, "label": label}
            if color:
                kwargs["color"] = color
            ax.bar(offsets, y_vals, **kwargs)
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.grid(True, alpha=0.3, axis="y")
    elif chart_type == "scatter":
        for i, (y_vals, label) in enumerate(zip(y_data_sets, y_headers)):
            color = series_colors[i] if i < len(series_colors) else None
            skwargs: dict = {"label": label}
            if color:
                skwargs["color"] = color
            ax.scatter(x_pos, y_vals, **skwargs)
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.grid(True, alpha=0.3)
    elif chart_type == "pie":
        y_vals = y_data_sets[0] if y_data_sets else []
        pie_kwargs: dict = {"labels": x_labels, "autopct": "%1.1f%%", "startangle": 90}
        valid_pie_colors = [c for c in series_colors if c]
        if valid_pie_colors:
            pie_kwargs["colors"] = valid_pie_colors
        ax.pie(y_vals, **pie_kwargs)
        ax.axis("equal")
    else:  # line
        for i, (y_vals, label) in enumerate(zip(y_data_sets, y_headers)):
            color = series_colors[i] if i < len(series_colors) else None
            lkwargs: dict = {"marker": "o", "label": label}
            if color:
                lkwargs["color"] = color
            ax.plot(x_pos, y_vals, **lkwargs)
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.grid(True, alpha=0.3)

    if title:
        ax.set_title(title, fontsize=13, pad=10)
    if x_label and chart_type != "pie":
        ax.set_xlabel(x_label)
    if y_label and chart_type != "pie":
        ax.set_ylabel(y_label)
    if chart_type != "pie":
        if y_min is not None or y_max is not None:
            current_min, current_max = ax.get_ylim()
            ax.set_ylim(y_min if y_min is not None else current_min, y_max if y_max is not None else current_max)
        if y_tick_step is not None and y_tick_step > 0:
            tick_min, tick_max = ax.get_ylim()
            tick = tick_min
            ticks = []
            guard = 0
            while tick <= tick_max + (y_tick_step * 0.001) and guard < 1000:
                ticks.append(tick)
                tick += y_tick_step
                guard += 1
            if ticks:
                ax.set_yticks(ticks)
        import matplotlib.ticker as _ticker
        if y_tick_format == "comma":
            ax.yaxis.set_major_formatter(_ticker.FuncFormatter(lambda x, _: f"{int(round(x)):,}"))
        else:
            ax.ticklabel_format(style="plain", axis="y", useOffset=False)
    if len(y_indices) > 1:
        ax.legend()

    fig.tight_layout()

    save_dir = output_path.parent if output_path is not None else Path(sys.argv[0]).parent
    png_path = save_dir / f"chart_{ws.title}_{insert_cell}.png"
    fig.savefig(str(png_path), dpi=150, bbox_inches="tight")
    plt.close(fig)

    # Replace images anchored at the requested chart position. Keep unrelated images.
    try:
        insert_row, insert_col = coordinate_to_tuple(insert_cell)
        ws._images = [
            image for image in ws._images
            if not (
                hasattr(getattr(image, "anchor", None), "_from")
                and image.anchor._from.row + 1 == insert_row
                and image.anchor._from.col + 1 == insert_col
            )
        ]
    except Exception:
        pass

    img = XLImage(str(png_path))
    img.anchor = insert_cell
    ws.add_image(img)
    print(f"[edit_excel] chart: inserted {chart_type} chart image at {insert_cell}", file=sys.stderr)
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--plan", required=True)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    plan_path = Path(args.plan)

    plan: dict[str, Any] = json.loads(plan_path.read_text(encoding="utf-8"))
    wb = load_workbook_any(input_path)
    for _ws in wb.worksheets:
        non_empty = []
        for row in _ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    non_empty.append((cell.row, cell.column, cell.value))
                    if len(non_empty) >= 6:
                        break
            if len(non_empty) >= 6:
                break
        _sample = non_empty
        print(f"[edit_excel] loaded sheet '{_ws.title}': max_row={_ws.max_row} max_col={_ws.max_column} sample={_sample[:6]}", file=sys.stderr)
    if wb.calculation is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    total_sheets = len(wb.sheetnames)
    changed_sheets: set[str] = set()

    # sheetEdits
    for sheet_edit in plan.get("sheetEdits") or []:
        sheet_name = str(sheet_edit.get("sheetName") or "").strip()
        ws = resolve_sheet(wb, sheet_name)

        if ws is None:
            # "*" または未指定 → 全シート
            targets = wb.worksheets
        else:
            targets = [ws]

        for target_ws in targets:
            n = apply_sheet_edits(target_ws, sheet_edit)
            if n > 0:
                changed_sheets.add(target_ws.title)

    # formatEdits
    for fmt in plan.get("formatEdits") or []:
        sheet_name = str(fmt.get("sheetName") or "").strip()
        ws = resolve_sheet(wb, sheet_name)
        if ws is None:
            print(f"[edit_excel] formatEdits: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        n = apply_format_edits(ws, fmt)
        if n > 0:
            changed_sheets.add(ws.title)

    # borderEdits
    for be in plan.get("borderEdits") or []:
        sheet_name = str(be.get("sheetName") or "").strip()
        ws = resolve_sheet(wb, sheet_name)
        if ws is None:
            print(f"[edit_excel] borderEdits: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        n = apply_border_edits(ws, be)
        if n > 0:
            changed_sheets.add(ws.title)

    # copyRowColorEdits: 参照列の背景色を対象列の各行にコピー
    for crc in plan.get("copyRowColorEdits") or []:
        sheet_name = str(crc.get("sheetName") or "").strip()
        ws = resolve_sheet(wb, sheet_name)
        if ws is None:
            print(f"[edit_excel] copyRowColorEdits: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        target_col = str(crc.get("targetColumn") or "").strip()
        reference_col = str(crc.get("referenceColumn") or "").strip()
        start_row = int(crc.get("startRow") or 2)
        n = apply_copy_row_color(ws, target_col, reference_col, start_row)
        if n > 0:
            changed_sheets.add(ws.title)

    # chartEdits: openpyxl ネイティブチャートをシートに挿入
    for ce in plan.get("chartEdits") or []:
        sheet_name = str(ce.get("sheetName") or "").strip()
        ws = resolve_sheet(wb, sheet_name)
        if ws is None:
            print(f"[edit_excel] chartEdits: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        ok = apply_chart_edits(ws, ce, output_path)
        if ok:
            changed_sheets.add(ws.title)

    wb.save(str(output_path))

    print(json.dumps({"changedSheets": len(changed_sheets), "totalSheets": total_sheets}))


if __name__ == "__main__":
    main()
